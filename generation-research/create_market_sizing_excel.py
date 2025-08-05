import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

# Define the data
generations = ['Silent Generation', 'Baby Boomers', 'Generation X', 'Millennials', 'Generation Z', 'Generation Alpha']
years = [2025, 2026, 2027, 2028, 2029, 2030]

# Population data (in millions)
population_data = {
    'Silent Generation': 13.2,
    'Baby Boomers': 76.4,
    'Generation X': 65.0,
    'Millennials': 74.2,
    'Generation Z': 67.0,
    'Generation Alpha': 12.8
}

# Daily AI Usage % (constant across years based on current data)
daily_ai_usage = {
    'Silent Generation': 0.02,  # 2%
    'Baby Boomers': 0.11,       # 11%
    'Generation X': 0.19,       # 19%
    'Millennials': 0.24,        # 24%
    'Generation Z': 0.21,       # 21%
    'Generation Alpha': 0.33    # 33%
}

# AI Purchase Adoption % by year
ai_purchase_adoption = {
    'Silent Generation': [0.02, 0.03, 0.05, 0.08, 0.12, 0.16],
    'Baby Boomers': [0.05, 0.09, 0.15, 0.23, 0.32, 0.40],
    'Generation X': [0.08, 0.14, 0.22, 0.32, 0.42, 0.50],
    'Millennials': [0.15, 0.28, 0.45, 0.62, 0.75, 0.83],
    'Generation Z': [0.18, 0.35, 0.55, 0.72, 0.82, 0.87],
    'Generation Alpha': [0.12, 0.22, 0.38, 0.55, 0.70, 0.80]
}

# Annual Discretionary Spend (in USD)
annual_spend = {
    'Silent Generation': 1200,
    'Baby Boomers': 50000,
    'Generation X': 47000,
    'Millennials': 31256,
    'Generation Z': 940,
    'Generation Alpha': 572
}

# Create Excel writer
excel_file = '/Users/nicholaspate/Documents/agentic-commerce/generation-research/agentic_commerce_market_sizing_2025_2030.xlsx'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# Sheet 1: Executive Summary
exec_summary_data = []
for year_idx, year in enumerate(years):
    year_total = 0
    for gen in generations:
        pop = population_data[gen]
        daily_ai = daily_ai_usage[gen]
        purchase_pct = ai_purchase_adoption[gen][year_idx]
        spend = annual_spend[gen]
        market_size = pop * 1000000 * daily_ai * purchase_pct * spend / 1000000000  # in billions
        year_total += market_size
    exec_summary_data.append({
        'Year': year,
        'Total Market Size (Billions USD)': round(year_total, 1),
        'YoY Growth %': 0 if year_idx == 0 else round(((year_total / exec_summary_data[year_idx-1]['Total Market Size (Billions USD)']) - 1) * 100, 1)
    })

df_exec_summary = pd.DataFrame(exec_summary_data)
df_exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

# Sheet 2: Market Calculation Details
calc_data = []
for year_idx, year in enumerate(years):
    for gen in generations:
        pop = population_data[gen]
        daily_ai = daily_ai_usage[gen]
        purchase_pct = ai_purchase_adoption[gen][year_idx]
        spend = annual_spend[gen]
        market_size = pop * 1000000 * daily_ai * purchase_pct * spend / 1000000000  # in billions
        
        calc_data.append({
            'Year': year,
            'Generation': gen,
            'Population (Millions)': pop,
            'Daily AI Usage %': f"{daily_ai * 100:.0f}%",
            'AI Purchase Adoption %': f"{purchase_pct * 100:.0f}%",
            'Annual Spend per User (USD)': f"${spend:,}",
            'Market Size (Billions USD)': round(market_size, 3),
            'Formula': f"{pop}M × {daily_ai*100:.0f}% × {purchase_pct*100:.0f}% × ${spend:,}"
        })

df_calculations = pd.DataFrame(calc_data)
df_calculations.to_excel(writer, sheet_name='Market Calculations', index=False)

# Sheet 3: Generation Breakdown
gen_breakdown_data = []
for gen in generations:
    row_data = {'Generation': gen}
    for year_idx, year in enumerate(years):
        pop = population_data[gen]
        daily_ai = daily_ai_usage[gen]
        purchase_pct = ai_purchase_adoption[gen][year_idx]
        spend = annual_spend[gen]
        market_size = pop * 1000000 * daily_ai * purchase_pct * spend / 1000000000
        row_data[f'{year}'] = round(market_size, 2)
    gen_breakdown_data.append(row_data)

# Add total row
total_row = {'Generation': 'TOTAL'}
for year in years:
    year_total = sum([row[f'{year}'] for row in gen_breakdown_data])
    total_row[f'{year}'] = round(year_total, 1)
gen_breakdown_data.append(total_row)

df_gen_breakdown = pd.DataFrame(gen_breakdown_data)
df_gen_breakdown.to_excel(writer, sheet_name='Generation Breakdown', index=False)

# Sheet 4: Input Parameters
params_data = []

# Population parameters
for gen in generations:
    params_data.append({
        'Category': 'Population (Millions)',
        'Generation': gen,
        'Value': population_data[gen],
        'Source': 'US Census Bureau / Statista / Pew Research',
        'URL': 'https://www.statista.com/statistics/797321/us-population-by-generation/'
    })

# Daily AI Usage parameters
for gen in generations:
    source_url = 'https://menlovc.com/perspective/2025-the-state-of-consumer-ai/'
    if gen == 'Generation Alpha':
        source_url = 'https://digitalwellnesslab.org/articles/key-takeaways-from-our-2024-pulse-survey-on-teens-and-ai/'
    elif gen == 'Silent Generation':
        source_url = 'Estimated from caregiver-mediated usage'
    
    params_data.append({
        'Category': 'Daily AI Usage %',
        'Generation': gen,
        'Value': f"{daily_ai_usage[gen] * 100:.0f}%",
        'Source': 'Menlo Ventures 2025' if gen not in ['Generation Alpha', 'Silent Generation'] else 'Digital Wellness Lab 2024' if gen == 'Generation Alpha' else 'Estimated',
        'URL': source_url
    })

# Annual Spend parameters
spend_sources = {
    'Silent Generation': ('Senior e-commerce patterns', 'https://www.outerboxdesign.com/digital-marketing/senior-citizen-ecommerce-habits'),
    'Baby Boomers': ('Assumption based on high-income segments', 'N/A'),
    'Generation X': ('Nielsen IQ', 'https://nielseniq.com/global/en/news-center/2025/overlooked-and-under-marketed-gen-x-emerges-as-most-influential-global-consumer-cohort/'),
    'Millennials': ('Capital One Shopping', 'https://capitaloneshopping.com/research/millennial-shopping-statistics/'),
    'Generation Z': ('Bango Subscription Study', 'https://www.thewrap.com/gen-z-most-subscribed-generation-study-bango/'),
    'Generation Alpha': ('Generation Alpha spending data', 'https://dealaid.org/data/gen-alpha/')
}

for gen in generations:
    params_data.append({
        'Category': 'Annual Spend (USD)',
        'Generation': gen,
        'Value': f"${annual_spend[gen]:,}",
        'Source': spend_sources[gen][0],
        'URL': spend_sources[gen][1]
    })

df_params = pd.DataFrame(params_data)
df_params.to_excel(writer, sheet_name='Input Parameters', index=False)

# Sheet 5: AI Purchase Adoption Rates
adoption_data = []
for gen in generations:
    row_data = {'Generation': gen}
    for year_idx, year in enumerate(years):
        row_data[f'{year}'] = f"{ai_purchase_adoption[gen][year_idx] * 100:.0f}%"
    adoption_data.append(row_data)

df_adoption = pd.DataFrame(adoption_data)
df_adoption.to_excel(writer, sheet_name='AI Purchase Adoption Rates', index=False)

# Sheet 6: Assumptions & Methodology
assumptions = [
    {
        'Category': 'Market Definition',
        'Assumption': 'Agentic Commerce = AI agents making autonomous purchasing decisions on behalf of users',
        'Rationale': 'Focus on truly autonomous AI purchasing, not just AI-assisted shopping'
    },
    {
        'Category': 'Population Data',
        'Assumption': 'Population remains constant 2025-2030',
        'Rationale': 'Simplification; actual population changes would be minimal over 5-year period'
    },
    {
        'Category': 'Daily AI Usage',
        'Assumption': 'Daily AI usage rates remain constant at 2025 levels',
        'Rationale': 'Conservative approach; actual usage likely to increase but offset by market saturation'
    },
    {
        'Category': 'Purchase Adoption Growth',
        'Assumption': 'S-curve adoption pattern with accelerating then decelerating growth',
        'Rationale': 'Follows typical technology adoption lifecycle patterns'
    },
    {
        'Category': 'Baby Boomer Spend',
        'Assumption': '$50,000 annual agentic commerce spend per adopter',
        'Rationale': 'High-income segments most likely to adopt; includes travel, healthcare, financial services'
    },
    {
        'Category': 'Generation Z Spend',
        'Assumption': '$940 based on subscription spending only',
        'Rationale': 'Conservative estimate; actual spending likely higher when including all categories'
    },
    {
        'Category': 'Trust-to-Purchase Conversion',
        'Assumption': 'AI purchase adoption lags behind general AI usage by 2-3 years',
        'Rationale': 'Financial transactions require higher trust levels than informational AI use'
    },
    {
        'Category': 'Market Coverage',
        'Assumption': 'US consumer market only',
        'Rationale': 'Focus on most mature market; global opportunity significantly larger'
    },
    {
        'Category': 'Age Cohort Stability',
        'Assumption': 'Generation Alpha limited to ages 13-15 throughout period',
        'Rationale': 'Maintains consistency; younger cohorts excluded due to limited purchasing power'
    },
    {
        'Category': 'Infrastructure Readiness',
        'Assumption': 'Technical and regulatory infrastructure will support projected growth',
        'Rationale': 'Major tech companies investing heavily in agentic commerce capabilities'
    }
]

df_assumptions = pd.DataFrame(assumptions)
df_assumptions.to_excel(writer, sheet_name='Assumptions', index=False)

# Sheet 7: Data Sources
sources_data = [
    {
        'Data Type': 'Population Statistics',
        'Source': 'US Census Bureau',
        'URL': 'https://www.census.gov/library/stories/2019/12/by-2030-all-baby-boomers-will-be-age-65-or-older.html',
        'Date': '2024'
    },
    {
        'Data Type': 'Population by Generation',
        'Source': 'Statista',
        'URL': 'https://www.statista.com/statistics/797321/us-population-by-generation/',
        'Date': '2024'
    },
    {
        'Data Type': 'Daily AI Usage',
        'Source': 'Menlo Ventures Consumer AI Survey',
        'URL': 'https://menlovc.com/perspective/2025-the-state-of-consumer-ai/',
        'Date': 'January 2025'
    },
    {
        'Data Type': 'Teen AI Usage',
        'Source': 'Digital Wellness Lab',
        'URL': 'https://digitalwellnesslab.org/articles/key-takeaways-from-our-2024-pulse-survey-on-teens-and-ai/',
        'Date': '2024'
    },
    {
        'Data Type': 'ChatGPT Usage',
        'Source': 'Pew Research Center',
        'URL': 'https://www.pewresearch.org/short-reads/2025/06/25/34-of-us-adults-have-used-chatgpt-about-double-the-share-in-2023/',
        'Date': 'June 2025'
    },
    {
        'Data Type': 'Gen X Spending',
        'Source': 'Nielsen IQ',
        'URL': 'https://nielseniq.com/global/en/news-center/2025/overlooked-and-under-marketed-gen-x-emerges-as-most-influential-global-consumer-cohort/',
        'Date': '2025'
    },
    {
        'Data Type': 'Millennial Spending',
        'Source': 'Capital One Shopping',
        'URL': 'https://capitaloneshopping.com/research/millennial-shopping-statistics/',
        'Date': '2024'
    },
    {
        'Data Type': 'Gen Z Subscriptions',
        'Source': 'Bango Study',
        'URL': 'https://www.thewrap.com/gen-z-most-subscribed-generation-study-bango/',
        'Date': '2024'
    },
    {
        'Data Type': 'Gen Alpha Spending',
        'Source': 'DealAid Research',
        'URL': 'https://dealaid.org/data/gen-alpha/',
        'Date': '2024'
    },
    {
        'Data Type': 'Senior E-commerce',
        'Source': 'OuterBox Design',
        'URL': 'https://www.outerboxdesign.com/digital-marketing/senior-citizen-ecommerce-habits',
        'Date': '2024'
    }
]

df_sources = pd.DataFrame(sources_data)
df_sources.to_excel(writer, sheet_name='Data Sources', index=False)

# Save and close the Excel file
writer.close()

print(f"Excel file created successfully: {excel_file}")

# Now apply formatting using openpyxl
wb = openpyxl.load_workbook(excel_file)

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
currency_format = '#,##0.0'
percent_format = '0%'

# Format Executive Summary
ws = wb['Executive Summary']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in row:
        cell.number_format = '$#,##0.0"B"'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        cell.number_format = '0.0"%"'

# Format Market Calculations
ws = wb['Market Calculations']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Format Generation Breakdown
ws = wb['Generation Breakdown']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Format the TOTAL row
for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = total_fill

# Format numbers in Generation Breakdown
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2):
    for cell in row:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.0"B"'

# Format Input Parameters
ws = wb['Input Parameters']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Format AI Purchase Adoption Rates
ws = wb['AI Purchase Adoption Rates']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Format Assumptions
ws = wb['Assumptions']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
# Wrap text for assumption cells
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Adjust column widths for Assumptions
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 60

# Format Data Sources
ws = wb['Data Sources']
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 80
ws.column_dimensions['D'].width = 15

# Add a chart to Executive Summary
ws = wb['Executive Summary']
chart = LineChart()
chart.title = "US Agentic Commerce Market Size Projection 2025-2030"
chart.y_axis.title = "Market Size (Billions USD)"
chart.x_axis.title = "Year"
chart.height = 10
chart.width = 15

# Add data to chart
data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=2)
years_ref = Reference(ws, min_col=1, min_row=2, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(years_ref)

ws.add_chart(chart, "E2")

# Save the formatted workbook
wb.save(excel_file)
print(f"Formatting applied successfully to: {excel_file}")